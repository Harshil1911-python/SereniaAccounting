"""
SERENIA ACCOUNTING — reports/generators.py
=============================================
PDF (ReportLab) and Excel (openpyxl) generators for financial reports.
Returns Django HttpResponse with appropriate content-type for download.
"""

from io import BytesIO
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill


class PDFReportGenerator:
    """Generates print-ready PDF reports using ReportLab."""

    PRIMARY_COLOR = colors.HexColor('#6C5CE7')
    HEADER_BG = colors.HexColor('#F0EEFF')

    @classmethod
    def trial_balance(cls, data: dict) -> HttpResponse:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=20 * mm, bottomMargin=20 * mm)
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('Title', parent=styles['Heading1'], textColor=cls.PRIMARY_COLOR, fontSize=18)
        elements = [
            Paragraph(data['company'], title_style),
            Paragraph(f"Trial Balance — {data['financial_year']}", styles['Heading3']),
            Paragraph(f"As of {data['as_of_date']}", styles['Normal']),
            Spacer(1, 12),
        ]

        table_data = [['Ledger', 'Group', 'Debit', 'Credit']]
        for row in data['rows']:
            table_data.append([
                row['ledger_name'], row['group'],
                row['balance'] if row['balance_type'] == 'Dr' else '',
                row['balance'] if row['balance_type'] == 'Cr' else '',
            ])
        table_data.append(['', 'Total', data['totals']['debit'], data['totals']['credit']])

        table = Table(table_data, colWidths=[150, 130, 90, 90])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), cls.HEADER_BG),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('LINEBELOW', (0, 0), (-1, 0), 1, cls.PRIMARY_COLOR),
            ('LINEABOVE', (0, -1), (-1, -1), 1, colors.black),
            ('ALIGN', (2, 0), (-1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E9EAEF')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAFAFA')]),
        ]))
        elements.append(table)

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="trial_balance_{data["as_of_date"]}.pdf"'
        return response


class ExcelReportGenerator:
    """Generates Excel (.xlsx) reports using openpyxl."""

    HEADER_FILL = PatternFill(start_color='6C5CE7', end_color='6C5CE7', fill_type='solid')
    HEADER_FONT = Font(color='FFFFFF', bold=True)

    @classmethod
    def trial_balance(cls, data: dict) -> HttpResponse:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Trial Balance'

        ws['A1'] = data['company']
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Trial Balance — {data['financial_year']}"
        ws['A3'] = f"As of {data['as_of_date']}"
        ws.append([])

        headers = ['Ledger', 'Group', 'Debit', 'Credit']
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = cls.HEADER_FILL
            cell.font = cls.HEADER_FONT
            cell.alignment = Alignment(horizontal='center')

        for row in data['rows']:
            ws.append([
                row['ledger_name'], row['group'],
                float(row['balance']) if row['balance_type'] == 'Dr' else None,
                float(row['balance']) if row['balance_type'] == 'Cr' else None,
            ])

        ws.append(['', 'Total', float(data['totals']['debit']), float(data['totals']['credit'])])
        total_row = ws.max_row
        for col_idx in range(1, 5):
            ws.cell(row=total_row, column=col_idx).font = Font(bold=True)

        # Column widths
        for col, width in zip('ABCD', [35, 25, 15, 15]):
            ws.column_dimensions[col].width = width

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="trial_balance_{data["as_of_date"]}.xlsx"'
        return response
